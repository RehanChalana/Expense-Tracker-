from flask import Flask, request, send_file
from flask_cors import CORS
import json
from reportlab.pdfgen import canvas
import openpyxl
import csv  
from datetime import datetime

app = Flask(__name__)
CORS(app)

def generate_pdf(data):
    # Your PDF generation logic here...
    # Create a PDF document 
    name=data["userData"][0]["walletname"]
    budget=data["userData"][0]["Budget"]
    userBalance=data["userData"][0]["userBalance"]
    pdf_filename = "expense_report.pdf"
    c = canvas.Canvas(pdf_filename)
    y=750
    c.setFont("Helvetica",25)
    c.drawString(50,820,f"Expense Report! For {name}'s Wallet")
    c.drawString(30,790,f"Balance:{userBalance}")
    c.drawString(220,790,f"Budget:{budget}")
    c.drawString(90,750,f"Title")
    c.drawString(240,750,f"Amount")
    c.drawString(340,750,f"Date")
    c.drawString(440,750,f"Category")
    c.setFont("Helvetica", 12)
    
    y_position = 730  # Initial Y position for expenses
    # x=100
    for i in data["historyData"]:
        Title = i["Title"]
        amount = i["Amount"]
        Date = i["Date"]
        category = i["category"]
        # Add expense details to the PDF
        c.drawString(100,y_position, f"{Title}")
        c.drawString(250,y_position,f"{amount}")
        c.drawString(350,y_position,f"{Date}")
        c.drawString(450,y_position,f"{category}")
        y_position -= 20  # Move to the next line
    c.save()

    return pdf_filename

def generate_xl(data):
       # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ["Title", "Amount", "Date", "Category"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data to the worksheet
    for row_num, expense in enumerate(data["historyData"], 2):
        ws.cell(row=row_num, column=1, value=expense["Title"])
        ws.cell(row=row_num, column=2, value=expense["Amount"])
        ws.cell(row=row_num, column=3, value=expense["Date"])
        ws.cell(row=row_num, column=4, value=expense["category"])

    # Save the workbook to a file
    xl_filename = "expense_report.xlsx"
    wb.save(xl_filename)

    return xl_filename

def generate_csv(data):
    # Your CSV generation logic here...
    # Create a CSV file
    
    csv_filename = "expense_report.csv"

    with open(csv_filename, 'w', newline='') as csvfile:
        fieldnames = ["Title", "Amount", "Date", "category"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        # Write headers to the CSV file
        writer.writeheader()

        # Write data to the CSV file
        for expense in data["historyData"]:
        # Convert the date string to a datetime object
            date_str = expense["Date"]
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            expense["Date"] = date_obj.strftime("%Y-%m-%d")

            writer.writerow(expense)

    return csv_filename
    

@app.route('/generate_pdf', methods=['POST'])
def generate_pdf_endpoint():
    data = request.get_json()  # Retrieve JSON data from the request
    if(data["request"]=="pdf"):
        file = generate_pdf(data)
        
    if(data["request"]=="xl"):
        file = generate_xl(data)
        
    if(data["request"]=="csv"):
        file = generate_csv(data)
    
    return send_file(file, as_attachment=True)

if __name__ == '__main__':
    app.run(port=5000)
